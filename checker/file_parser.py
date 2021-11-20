import string
from datetime import datetime
from multiprocessing.pool import ThreadPool

from openpyxl import load_workbook

from models import SubTask, Product, SubTaskHasCategory, Category, ProductHasCategory
from name_checker import neuron_oracle

product_name_column_name = 'Общее наименование продукции'
product_code_column_name = 'Раздел ЕП РФ (Код из ФГИС ФСА для подкатегории продукции)'


def handle_product(db, task_id, name: str, cat_codes: str):
    product = Product(name=name)
    db.session.add(product)
    db.session.commit()
    sub_task = SubTask(task_id=task_id, product_id=product.id, is_valid=False)
    db.session.add(sub_task)
    db.session.commit()
    user_codes = []
    user_categories = []
    for cat_code in cat_codes.split(';'):
        user_codes.append(cat_code.strip())
        category = db.session.query(Category).filter(Category.code == cat_code).first()
        if category:
            user_categories.append(category)
            stc = SubTaskHasCategory(sub_task_id=sub_task.id, category_code=category.code)
            db.session.add(stc)
            db.session.commit()

    our_categories, is_valid = neuron_oracle(name, user_codes)
    for category in our_categories:
        phc = ProductHasCategory(product_id=product.id, category_code=category.code)
        db.session.add(phc)
    sub_task.is_valid = is_valid
    db.session.commit()
    db.session.flush()


def file_to_farsh(db, filename, file_ext, task):
    if file_ext == 'xlsx':
        start_time = datetime.now()
        xlsx = load_workbook(filename)
        tab = xlsx.worksheets[0]
        product_name_column = None
        product_code_column = None

        for sym in string.ascii_uppercase:
            cell = tab[f'{sym}1'].value
            print(cell)
            if cell and product_name_column_name in cell:
                print('product_name_column', cell)
                product_name_column = sym
            if cell and product_code_column_name in cell:
                print('product_code_column', cell)
                product_code_column = sym
            if product_name_column and product_code_column:
                break

        if not product_name_column or not product_code_column:
            raise RuntimeError(f'product_name_column={product_name_column} product_code_column={product_code_column}')

        db.session.add(task)
        db.session.commit()
        pool = ThreadPool(processes=16)

        i = 2
        while True:
            name = tab[f'{product_name_column}{i}'].value
            cat_codes = tab[f'{product_code_column}{i}'].value
            if name:
                pool.apply_async(handle_product, (db, task.id, name.strip(), cat_codes.strip()))
                i += 1
            else:
                print('doc end')
                break
        pool.close()
        pool.join()
        print(f"Thread handle {i-1} products for {datetime.now() - start_time}")
        task.status = 'done'
        db.session.commit()
        db.session.flush()
    else:
        raise NotImplementedError('xlsx only plz')


def normalize_str(text: str) -> str:
    return text.replace('\n', ' ').replace('\t', ' ').replace("\xa0", " ").strip()


def extract_categories_from_file(db, filename):
    try:
        print(f'Add categories from {filename} to database.\nPlease wait')
        xlsx = load_workbook(filename)
        tab = xlsx.worksheets[1]
        category_name_column = None
        category_code_column = None
        subcategory_name_column = None
        subcategory_code_column = None

        for sym in string.ascii_uppercase:
            cell = tab[f'{sym}1'].value
            print(cell)
            if cell and 'Код из ФГИС ФСА для подкатегории продукции' in cell:
                subcategory_code_column = sym
            if cell and 'Подкатегория продукции' in cell:
                subcategory_name_column = sym
            if cell and 'Категория продукции' in cell:
                category_name_column = sym
            if cell and 'Раздел ЕП РФ (Код)' in cell:
                category_code_column = sym
            if subcategory_name_column and subcategory_code_column:
                break

        if not subcategory_name_column or not subcategory_code_column:
            raise RuntimeError(f'subcategory_name_column={subcategory_name_column} '
                               f'subcategory_code_column={subcategory_code_column}')

        categories = []
        sub_categories = []
        homeless_sub_categories = []

        i = 2
        while True:

            if category_name_column and category_code_column:
                cat_name = tab[f'{category_name_column}{i}'].value
                cat_code = tab[f'{category_code_column}{i}'].value
                print(i, 1, cat_code, normalize_str(cat_name))
                if cat_name:
                    print(i, 2, cat_code, normalize_str(cat_name))
                    categories.append(Category(code=cat_code, name=normalize_str(cat_name)))
                    i += 1
                else:
                    print('doc end')
                    break
            else:
                cat_code = None
                cat_name = None

            if subcategory_name_column and subcategory_code_column:
                sub_cat_name = tab[f'{subcategory_name_column}{i}'].value
                sub_cat_code = tab[f'{subcategory_code_column}{i}'].value

                if sub_cat_name:
                    sub_cat_code = str(sub_cat_code).strip()

                    if '.' in sub_cat_code:
                        cat_code, _ = sub_cat_code.split('.', maxsplit=1)

                    if cat_code:
                        homeless_sub_categories.append(Category(code=cat_code,
                                                                name=normalize_str(sub_cat_name)))
                        sub_categories.append(Category(code=sub_cat_code,
                                                       name=normalize_str(sub_cat_name),
                                                       parent_code=cat_code))
                    else:
                        categories.append(Category(code=sub_cat_code, name=normalize_str(sub_cat_name)))

                    if not cat_name:
                        i += 1
                else:
                    print('doc end')
                    break

        db.session.bulk_save_objects(categories)
        db.session.commit()
        db.session.bulk_save_objects(homeless_sub_categories)
        db.session.commit()
        db.session.bulk_save_objects(sub_categories)
        db.session.commit()
    except Exception as e:
        import traceback
        print(e)
        traceback.print_exc()
